import openpyxl
import sys
from googleTranslator import Google
def main():
    fileName = 'Ranking30.xlsx'
    # Select which lines of the input sentences you wish to use
    input_selection = [1, 18975]
    googleTranslator = Google()
    try:
        file = openpyxl.load_workbook(fileName)
    except:
        file = openpyxl.Workbook()

    sheet = file['Sheet1']

    for line_counter in range(input_selection[0], input_selection[1]):
        try:
            sentence = sheet.cell(row=line_counter, column=7).value
            translations = googleTranslator.translate(sentence, 'en')
            sheet.cell(row = line_counter, column = 8).value = translations['translatedText']
        except Exception as exception:
            print (exception)
            continue
    # Save the file and notify the user
    file.save(fileName)
    print('Translations have been completed')


if __name__ == "__main__":
    main()
