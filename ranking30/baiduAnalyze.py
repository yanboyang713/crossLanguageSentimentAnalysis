import openpyxl
import sys
from baidu import Baidu
def main():
    fileName = 'Ranking30.xlsx'
    # Select which lines of the input sentences you wish to use
    input_selection = [1, 18975]
    baidu = Baidu()
    try:
        file = openpyxl.load_workbook(fileName)
    except:
        file = openpyxl.Workbook()

    sheet = file['Sheet1']

    for line_counter in range(input_selection[0], input_selection[1]):
        try:
            sentence = sheet.cell(row=line_counter, column=7).value
            baidu.analyze(sentence)
            sheet.cell(row = line_counter, column = 9).value = baidu.getPositiveProb()
            sheet.cell(row = line_counter, column = 10).value = baidu.getConfidence()
            sheet.cell(row = line_counter, column = 11).value = baidu.getNegativeProb()
            sheet.cell(row = line_counter, column = 12).value = baidu.getSentiment()
        except Exception as exception:
            print (exception)
            continue
    # Save the file and notify the user
    file.save(fileName)
    print("baidu analyze finish")


if __name__ == "__main__":
    main()
