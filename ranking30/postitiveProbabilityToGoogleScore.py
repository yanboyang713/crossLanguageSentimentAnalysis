import openpyxl
import sys
from decimal import *
def main():
    fileName = 'Ranking30.xlsx'
    # Select which lines of the input sentences you wish to use
    input_selection = [1, 18975]
    input_column = 9
    output_column = 31
    try:
        file = openpyxl.load_workbook(fileName)
    except:
        file = openpyxl.Workbook()

    sheet = file['Sheet1']

    for line_counter in range(input_selection[0], input_selection[1]):
        try:
            inputValue = sheet.cell(row=line_counter, column=input_column).value
            if inputValue:
                inputValue = Decimal(inputValue)
                getcontext().prec = 30
                if (inputValue < 0.5):
                    result = (1 - inputValue) * -1
                elif (inputValue > 0.5):
                    result = inputValue
                else:
                    result = 0
                sheet.cell(row = line_counter, column = output_column).value = Decimal(result)
        except Exception as exception:
            print (exception)
            continue
    # Save the file and notify the user
    file.save(fileName)
    print("change finish")


if __name__ == "__main__":
    main()
