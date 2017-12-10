import openpyxl
import sys
import json
from yandex_translate import YandexTranslate
import string
import ast

def main():
    fileName = 'Ranking30.xlsx'
    translate = YandexTranslate('trnsl.1.1.20171209T032531Z.b7d8027c269f0075.dc087b58baea72d8d0ec2dd4843879cebdd41c1f')

    # Select which lines of the input sentences you wish to use
    input_selection = [14000, 18975]
    try:
        file = openpyxl.load_workbook(fileName)
    except:
        file = openpyxl.Workbook()

    sheet = file['Sheet1']

    for line_counter in range(input_selection[0], input_selection[1]):
        try:
            sentence = sheet.cell(row=line_counter, column=7).value
            translations = json.dumps(ast.literal_eval(str(translate.translate(sentence, 'zh-en'))))

            jsonObj = json.loads(translations)
            sheet.cell(row = line_counter, column = 21).value = jsonObj["text"][0]
        except Exception as exception:
            print (exception)
            if exception == "ERR_DAILY_CHAR_LIMIT_EXCEDED":
                file.save(fileName)
                return
            continue
    # Save the file and notify the user
    file.save(fileName)
    print('Translations have been completed')


if __name__ == "__main__":
    main()
