import openpyxl
import sys
from bing import Bing
def main():
    fileName = 'Ranking0.xlsx'
    # Select which lines of the input sentences you wish to use
    input_selection = [1, 3]
    bingTranslator = Bing()
    try:
        file = openpyxl.load_workbook(fileName)
    except:
        file = openpyxl.Workbook()

    sheet = file['Sheet1']

    for line_counter in range(input_selection[0], input_selection[1]):
        try:
            sentence = sheet.cell(row=line_counter, column=7).value
            translations = bingTranslator.translate(sentence, 'en', 'zh-CHS')
            print (translations)
            #sheet.cell(row = line_counter, column = 9).value = translations
        except Exception as exception:
            print (exception)
            continue
    # Save the file and notify the user
    file.save(fileName)
    print('Translations have been completed')


if __name__ == "__main__":
    main()
