import openpyxl
def main():
    fileName = 'Ranking10.xlsx'
    # Select which lines of the input sentences you wish to use
    input_selection = [1, 8574]
    try:
        file = openpyxl.load_workbook(fileName)
    except:
        file = openpyxl.Workbook()

    sheet = file['Sheet1']

    for line_counter in range(input_selection[0], input_selection[1]):
        try:
            sentence = sheet.cell(row=line_counter, column = 7).value
            for index in range(line_counter + 1, input_selection[1]):
                compare = sheet.cell(row = index, column = 7).value
                if sentence == compare and sentence != "":
                    sheet.cell(row = index, column = 7).value = ""
                    print ("error", line_counter, " = ", sentence, index, " = ", compare)
        except Exception as exception:
            print (exception)
            continue
    # Save the file and notify the user
    file.save(fileName)
    print("remove Redundancy finish")


if __name__ == "__main__":
    main()
