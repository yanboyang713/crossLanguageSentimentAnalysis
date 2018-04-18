import numpy as np
from openpyxl import load_workbook
import matplotlib as mpl

## agg backend is used to create plot as a .png file
mpl.use('agg')

import matplotlib.pyplot as plt

def drawBoxPlots(title, data_to_plot, ax):

    # Create the boxplot
    bp = ax.boxplot(data_to_plot, patch_artist=True, showmeans=True)
    ## change outline color, fill color and linewidth of the boxes
    for box in bp['boxes']:
        # change outline color
        box.set( color='#7570b3', linewidth=2)
        # change fill color
        box.set( facecolor = '#1b9e77' )

    ## change color and linewidth of the whiskers
    for whisker in bp['whiskers']:
        whisker.set(color='#7570b3', linewidth=2)

    ## change color and linewidth of the caps
    for cap in bp['caps']:
        cap.set(color='#7570b3', linewidth=2)

    ## change color and linewidth of the medians
    for median in bp['medians']:
        median.set(color='#b2df8a', linewidth=2)

    ## change the style of fliers and their fill
    for flier in bp['fliers']:
        flier.set(marker='o', markerfacecolor='red', markersize=5, markeredgewidth=0.0, alpha=0.5)

    for mean in bp['means']:
        mean.set(marker = 's', markerfacecolor='red')

    ## Custom x-axis labels
    ax.set_xticklabels(['Ranking10', 'Ranking20', 'Ranking30', 'Ranking40', 'Ranking50'])
    ax.set_title(title)


## Create data

ranking10 = list()
ranking20 = list()
ranking30 = list()
ranking40 = list()
ranking50 = list()
wb = load_workbook(filename='good.xlsx', read_only=True)
ws = wb['Sheet1']

for row in range(1, 46181):
    ranking = ws.cell(row=row, column=6).value
    value = float(ws.cell(row=row, column=17).value)
    if ranking == 10:
        print ("10")
        ranking10.append(value)
    elif ranking == 20:
        print ("20")
        ranking20.append (value)
    elif ranking == 30:
        print ("30")
        ranking30.append (value)
    elif ranking == 40:
        print ("40")
        ranking40.append (value)
    elif ranking == 50:
        print ("50")
        ranking50.append (value)
## combine these different collections into a list
print ("get data")
data_to_plot = [ranking10, ranking20, ranking30, ranking40, ranking50]


#fig, axes = plt.subplots(nrows=2, ncols=4, figsize=(9, 4))
# Create a figure instance
fig = plt.figure(1, figsize=(9, 6))
# Create an axes instance
ax = fig.add_subplot(111)
## add patch_artist=True option to ax.boxplot()
## to get fill color


drawBoxPlots("google chinese sentiment analysis", data_to_plot, ax)
# Save the figure
fig.savefig("googleChineseSentimentAnalysis.png", bbox_inches='tight')
