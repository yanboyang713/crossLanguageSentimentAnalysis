import openpyxl
class classift(object):
    def __init__(self):
        self.ranking10FileName = "Ranking10.xlsx"
        self.ranking20FileName = "Ranking20.xlsx"
        self.ranking30FileName = "Ranking30.xlsx"
        self.ranking40FileName = "Ranking40.xlsx"
        self.ranking50FileName = "Ranking50.xlsx"

        self.goodFileName = "good.xlsx"
        self.badFileName = "bad.xlsx"

        self.ranking10File = openpyxl.load_workbook(filename=self.ranking10FileName, read_only=True)
        self.ranking20File = openpyxl.load_workbook(filename=self.ranking20FileName, read_only=True)
        self.ranking30File = openpyxl.load_workbook(filename=self.ranking30FileName, read_only=True)
        self.ranking40File = openpyxl.load_workbook(filename=self.ranking40FileName, read_only=True)
        self.ranking50File = openpyxl.load_workbook(filename=self.ranking50FileName, read_only=True)

        self.goodFile = openpyxl.Workbook()
        self.badFile = openpyxl.Workbook()

        self.ranking10Sheet = self.ranking10File['Sheet1']
        self.ranking20Sheet = self.ranking20File['Sheet1']
        self.ranking30Sheet = self.ranking30File['Sheet1']
        self.ranking40Sheet = self.ranking40File['Sheet1']
        self.ranking50Sheet = self.ranking50File['Sheet1']

        self.goodSheet = self.goodFile.create_sheet()
        self.badSheet = self.badFile.create_sheet()

        self.goodSize = int(input ("Please input good file index for start: "))
        self.badSize = int (input ("Please input bad file index for start: "))

        self.run()


    def run(self):

        try:
            self.readFile(self.ranking10Sheet, 1, 8573)
            #self.readFile(self.ranking20Sheet, 1, 13227)
            #self.readFile(self.ranking30Sheet, 1, 18975)
            #self.readFile(self.ranking40Sheet, 1, 8791)
            #self.readFile(self.ranking50Sheet, 1, 4308)

        except Exception as exception:
            print (exception)

        # Save the file and notify the user
        self.goodFile.save(self.goodFileName)
        self.badFile.save(self.badFileName)
        print ("good file index: ", self.goodSize)
        print ("bad file index: ", self.badSize)
        print("finish")

    def readFile(self, sheetName, rowStart, rowEnd):
        for row in range(rowStart, rowEnd):
            flagNoEmpty = True
            for cell in range(6, 39):
                if sheetName.cell(row = row, column = cell).value == None:
                    self.writeFile(sheetName, row, "has empty")
                    flagNoEmpty = False
                    break
            if flagNoEmpty == True:
                self.writeFile(sheetName, row, "no empty")

    def writeFile(self, sheetName, row, which):
        if which == "has empty":
            for index in range(1, 39):
                self.badSheet.cell(row = self.badSize, column = index).value = sheetName.cell(row = row, column = index).value
            self.badSize += 1
        elif which == "no empty":
            for index in range(1, 39):
                self.goodSheet.cell(row = self.goodSize, column = index).value = sheetName.cell(row = row, column = index).value
            self.goodSize += 1

if __name__ == "__main__":
    classift()
