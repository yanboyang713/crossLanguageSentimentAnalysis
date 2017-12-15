from baiduTranslate import BaiduTranslate
import openpyxl
import sys
def main():
    fileName = 'Ranking0.xlsx'
    # Select which lines of the input sentences you wish to use
    input_selection = [1, 3]
    baidu = BaiduTranslate()
    try:
        file = openpyxl.load_workbook(fileName)
    except:
        file = openpyxl.Workbook()

    sheet = file['Sheet1']

    for line_counter in range(input_selection[0], input_selection[1]):
        try:
            sentence = sheet.cell(row=line_counter, column=7).value
            sheet.cell(row = line_counter, column = 24).value = baidu.translate("zh", "en", sentence)
        except Exception as exception:
            print (exception)
            continue
    # Save the file and notify the user
    file.save(fileName)
    print("baidu translate finish")


if __name__ == "__main__":
    main()
