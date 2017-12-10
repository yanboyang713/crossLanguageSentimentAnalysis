import openpyxl
import sys
import json
from yandex_translate import YandexTranslate
import string
import ast

def main():
    fileName = 'Ranking20.xlsx'
    translate = YandexTranslate('trnsl.1.1.20171207T010146Z.90e9c63da81d00aa.e54099dd331a9c656689519d90179363752428e7')

    # Select which lines of the input sentences you wish to use
    input_selection = [1, 13227]
    try:
        file = openpyxl.load_workbook(fileName)
    except:
        file = openpyxl.Workbook()

    sheet = file['Sheet1']

    for line_counter in range(input_selection[0], input_selection[1]):
        try:
            sentence = sheet.cell(row=line_counter, column=7).value
            translations = json.dumps(ast.literal_eval(str(translate.translate(sentence, 'zh-en'))))

            jsonObj = json.loads(result)
            sheet.cell(row = line_counter, column = 8).value = jsonObj["text"][0]
        except Exception as exception:
            print (exception)
            continue
    # Save the file and notify the user
    file.save(fileName)
    print('Translations have been completed')


if __name__ == "__main__":
    main()
