import openpyxl
import sys
from decimal import *
def main():
    fileName = 'Ranking10.xlsx'
    # Select which lines of the input sentences you wish to use
    input_selection = [1, 8573]
    input_column_one = 17
    input_column_two = 19
    try:
        file = openpyxl.load_workbook(fileName)
    except:
        file = openpyxl.Workbook()

    sheet = file['Sheet1']

    for line_counter in range(input_selection[0], input_selection[1]):
        try:
            inputValueOne = sheet.cell(row=line_counter, column=input_column_one).value
            inputValueTwo = sheet.cell(row=line_counter, column=input_column_two).value
            if inputValueOne is not None and inputValueTwo is not None:
                inputValueOne = Decimal(inputValueOne)
                inputValueTwo = Decimal(inputValueTwo)
                getcontext().prec = 30

                #print (inputValueOne, " ", inputValueTwo)
                if inputValueOne < 0 and inputValueTwo > 0:
                    print ("error! ", inputValueOne, " ", inputValueTwo)

                if inputValueOne > 0 and inputValueTwo < 0:
                    print ("error! ", inputValueOne, " ", inputValueTwo)

        except Exception as exception:
            print (exception)
            continue

    print("finish")


if __name__ == "__main__":
    main()
