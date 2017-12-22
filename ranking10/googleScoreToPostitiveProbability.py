import openpyxl
import sys
from decimal import *
def main():
    fileName = 'Ranking10.xlsx'
    # Select which lines of the input sentences you wish to use
    #input_selection = [1, 8573]
    input_selection = [1, 3]
    try:
        file = openpyxl.load_workbook(fileName)
    except:
        file = openpyxl.Workbook()

    sheet = file['Sheet1']

    for line_counter in range(input_selection[0], input_selection[1]):
        try:
            score = sheet.cell(row=line_counter, column=17).value
            getcontext().prec = 28
            result = (Decimal(score) + Decimal(1)) / Decimal(2)
            sheet.cell(row = line_counter, column = 31).value = result
        except Exception as exception:
            print (exception)
            continue
    # Save the file and notify the user
    file.save(fileName)
    print("change finish")


if __name__ == "__main__":
    main()
