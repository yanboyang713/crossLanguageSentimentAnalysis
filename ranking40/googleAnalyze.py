import openpyxl
import sys
from googleSentiment import Google
def main():
    fileName = 'Ranking40.xlsx'
    # Select which lines of the input sentences you wish to use
    input_selection = [1, 8791]
    google = Google()
    try:
        file = openpyxl.load_workbook(fileName)
    except:
        file = openpyxl.Workbook()

    sheet = file['Sheet1']

    for line_counter in range(input_selection[0], input_selection[1]):
        try:
            sentence = sheet.cell(row=line_counter, column=24).value
            sentence = sentence.encode()
            #zh for Chinese es for English
            languageIn = "es"
            google.analyze(sentence, languageIn)
            sheet.cell(row = line_counter, column = 25).value = google.getScore()
            sheet.cell(row = line_counter, column = 26).value = google.getMagnitude()
        except Exception as exception:
            print (exception)
            continue
    # Save the file and notify the user
    file.save(fileName)
    print("google analyze finish")

if __name__ == "__main__":
    main()
