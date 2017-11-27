import sys
import openpyxl
import pandas as pd
class separateFileByRank(object):
    def __init__(self):
        self.fileName = ""
        self.fromLine = 0
        self.toLine = 0
        self.whichRank = -1
        self.newFileRowCount = 1
        self.outputFileName = ""

    def main(self):
        self.fileName = input ("Please input file name: ")
        self.fromLine = input ("Which line do you want to start: ")
        self.toLine = input ("Which line do you want to end: ")
        self.whichRank = input ("Which ranking you want to separate: ")
        self.outputFileName = str(self.whichRank) + "Ranking.xlsx"

        self.createSpecialRankFile()

    def createSpecialRankFile(self):
        originWorkBook = openpyxl.load_workbook(self.fileName, read_only = True)
        originWorkingSheet = originWorkBook['Sheet1']

        newWorkBook = openpyxl.Workbook()
        newWorkSheet = newWorkBook.active

        try:
            for rowCounter in range (int(self.fromLine), int(self.toLine)):
                if int(self.whichRank) == originWorkingSheet.cell (row = rowCounter, column = 6).value:
                    #copy every column in current row
                    for columnCounter in range(1, 8):
                        if columnCounter == 2:
                            newWorkSheet.cell(row = self.newFileRowCount, column = columnCounter).value = pd.DateFrame(originWorkingSheet.cell(row = rowCounter, column = columnCounter).values)
                        newWorkSheet.cell(row = self.newFileRowCount, column = columnCounter).value = originWorkingSheet.cell(row = rowCounter, column = columnCounter).value

                    self.newFileRowCount = self.newFileRowCount + 1

        except Exception as exception:
            print (exception)

        finally:
            #save the file
            newWorkBook.save(self.outputFileName)
            print ("Completed !")
