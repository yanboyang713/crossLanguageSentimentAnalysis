import openpyxl
import sys
from baidu import Baidu
def main():
    fileName = 'Ranking50.xlsx'
    # Select which lines of the input sentences you wish to use
    input_selection = [1, 4308]
    baidu = Baidu()
    try:
        file = openpyxl.load_workbook(fileName)
    except:
        file = openpyxl.Workbook()

    sheet = file['Sheet1']

    for line_counter in range(input_selection[0], input_selection[1]):
        try:
            sentence = sheet.cell(row=line_counter, column=21).value
            baidu.analyze(sentence)
            sheet.cell(row = line_counter, column = 33).value = baidu.getPositiveProb()
            sheet.cell(row = line_counter, column = 34).value = baidu.getConfidence()
            sheet.cell(row = line_counter, column = 35).value = baidu.getNegativeProb()
            sheet.cell(row = line_counter, column = 36).value = baidu.getSentiment()
        except Exception as exception:
            print (exception)
            continue
    # Save the file and notify the user
    file.save(fileName)
    print("baidu analyze finish")


if __name__ == "__main__":
    main()
